import os
import requests
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime, timezone
import schedule
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# NOAA Weather API endpoint for searching alerts
ATOM_FEED_URL = "https://api.weather.gov/alerts/active.atom"

# States to monitor
STATES = sorted(["AL", "IL", "KS", "LA", "MO", "MS", "SC", "WI", "GA", "MI", "FL", "OH"])

# Fetch email credentials from environment variables
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

# Recipient email
RECIPIENT_EMAIL = "admin@mortgagebankersfs.com"

def fetch_weather_alerts():
    """Fetch current severe weather alerts for specified states using the Atom feed."""
    try:
        response = requests.get(ATOM_FEED_URL)
        response.raise_for_status()

        # Parse the Atom feed
        root = ET.fromstring(response.content)

        # Namespace for Atom feed
        ns = {'atom': 'http://www.w3.org/2005/Atom'}

        # Exclusion terms for the title
        exclusion_terms = [
            "Small Craft Advisory", "Gale Warning", "Gale Watch", "Open Water", 
            "Child Abduction Emergency", "AMBER Alert", "High Surf Advisory", 
            "High Surf Warning", "Rip Current", "Heavy Freezing Spray Warning", 
            "Spray Warning"
        ]

        # Extract relevant entries
        alerts = []
        for entry in root.findall("atom:entry", ns):
            title = entry.find("atom:title", ns).text or ""
            summary = entry.find("atom:summary", ns).text or ""
            updated = entry.find("atom:updated", ns).text or ""

            # Derive state from the last two characters of the title
            state = title[-2:] if title[-2:] in STATES else "Unknown"

            # Skip alerts for states not in the monitoring list
            if state == "Unknown":
                continue

            # Exclude alerts with specific terms in the title
            if any(term.lower() in title.lower() for term in exclusion_terms):
                continue

            # Format the updated timestamp
            try:
                updated_dt = datetime.fromisoformat(updated.replace("Z", "+00:00"))
                updated_formatted = updated_dt.strftime("%m/%d/%Y %H:%M-%H:%M")
            except ValueError:
                updated_formatted = updated

            # Parse the summary into WHAT, WHERE, WHEN, IMPACTS, ADDITIONAL DETAILS, INSTRUCTIONS
            details_split = {"WHAT": "", "WHERE": "", "WHEN": "", "IMPACTS": "", "ADDITIONAL DETAILS": "", "INSTRUCTIONS": ""}
            current_key = None
            for line in summary.split("\n"):
                line = line.strip()
                if line.startswith("* WHAT"):
                    current_key = "WHAT"
                    details_split[current_key] = line.replace("* WHAT", "").strip()
                elif line.startswith("* WHERE"):
                    current_key = "WHERE"
                    details_split[current_key] = line.replace("* WHERE", "").strip()
                elif line.startswith("* WHEN"):
                    current_key = "WHEN"
                    details_split[current_key] = line.replace("* WHEN", "").strip()
                elif line.startswith("* IMPACTS"):
                    current_key = "IMPACTS"
                    details_split[current_key] = line.replace("* IMPACTS", "").strip()
                elif line.startswith("* ADDITIONAL DETAILS"):
                    current_key = "ADDITIONAL DETAILS"
                    details_split[current_key] = line.replace("* ADDITIONAL DETAILS", "").strip()
                elif line.startswith("* INSTRUCTIONS"):
                    current_key = "INSTRUCTIONS"
                    details_split[current_key] = line.replace("* INSTRUCTIONS", "").strip()
                elif current_key:
                    details_split[current_key] += f" {line}"  # Append to the current key if it's a continuation

            alerts.append({
                "State": state,
                "Title": title,
                "Updated": updated_formatted,
                "WHAT": details_split["WHAT"],
                "WHERE": details_split["WHERE"],
                "WHEN": details_split["WHEN"],
                "IMPACTS": details_split["IMPACTS"],
                "ADDITIONAL DETAILS": details_split["ADDITIONAL DETAILS"],
                "INSTRUCTIONS": details_split["INSTRUCTIONS"]
            })

        print(f"Filtered down to {len(alerts)} alerts.")  # Debugging step
        return alerts

    except Exception as e:
        print(f"Error fetching weather alerts: {e}")
        return []

def save_to_spreadsheet(alerts):
    """Save the alerts to an Excel spreadsheet with today's date in the file name."""
    if not alerts:
        print("No alerts to save.")
        return None

    # Generate file name with today's date
    today = datetime.now().strftime("%Y-%m-%d")
    output_file = f"weather_alerts_{today}.xlsx"

    # Create a DataFrame
    df = pd.DataFrame(alerts, columns=["State", "Title", "Updated", "WHAT", "WHERE", "WHEN", "IMPACTS", "ADDITIONAL DETAILS", "INSTRUCTIONS"])

    # Sort by State and Updated columns
    df.sort_values(by=["State", "Updated"], inplace=True)

    # Save to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Adjust column widths and wrap text
    wb = load_workbook(output_file)
    ws = wb.active
    for col in ws.columns:
        column_letter = get_column_letter(col[0].column)
        ws.column_dimensions[column_letter].width = 42  # Set column width to 42
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping

    wb.save(output_file)
    print(f"Alerts saved to {output_file}")
    return output_file

def send_email_with_attachment(alert_message, num_alerts, attachment_path):
    """Send the alert message via email with the spreadsheet as an attachment."""
    try:
        current_date = datetime.now().strftime("%m/%d/%Y")
        subject = f"Severe Weather Alerts - {current_date} - {num_alerts} Alert(s) Found"
        msg = MIMEMultipart()
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = RECIPIENT_EMAIL
        msg["Subject"] = subject

        # Email body
        msg.attach(MIMEText(alert_message, "plain"))

        # Attach the spreadsheet
        if attachment_path:
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={os.path.basename(attachment_path)}",
            )
            msg.attach(part)

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, RECIPIENT_EMAIL, msg.as_string())
        print("Email sent successfully!")
    except Exception as e:
        print(f"Error sending email: {e}")

def daily_weather_alert():
    """Fetch alerts, save to a spreadsheet, and send via email."""
    alerts = fetch_weather_alerts()
    num_alerts = len(alerts)
    attachment_path = save_to_spreadsheet(alerts)
    alert_message = f"Attached is the weather alert report with {num_alerts} alert(s)."
    send_email_with_attachment(alert_message, num_alerts, attachment_path)

# Run the daily weather alert function once at startup
daily_weather_alert()

# Schedule the task daily at 6:30 AM CST (12:30 PM UTC)
schedule.every().day.at("06:30").do(daily_weather_alert)

print("Weather alert automation running...")
while True:
    schedule.run_pending()
    time.sleep(1)
